from datetime import datetime
import logging
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
import openpyxl
from ...utils import get_export_fullpath
from ...morphimport import wikimorph_import

log = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Import users from file'

    def add_arguments(self, parser):
        parser.add_argument('--filename', type=str, help='File to import')
        parser.add_argument('--start-row', type=int, help='Start row')
        parser.add_argument('--end-row', type=int, help='End row')

    def handle(self, *args, **options):
        start_time = datetime.now()
        filename = options.get('filename')
        start_row = options.get('start_row')
        end_row = options.get('end_row')
        file_path = get_export_fullpath(filename)
        log.info(f'Reading users from {file_path}')
        create_users_from_file(file_path, start_row, end_row)
        end_time = datetime.now()
        log.info(f'Import trajao ukupno {str(end_time-start_time)}')


def create_users_from_file(filename, start_row, end_row):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for i in range(start_row, end_row + 1):
        ime = sheet.cell(row=i, column=1).value
        prezime = sheet.cell(row=i, column=2).value
        email = sheet.cell(row=i, column=3).value
        lozinka = sheet.cell(row=i, column=4).value
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            user = User.objects.create_user(email, email=email, password=lozinka)
            user.first_name = ime
            user.last_name = prezime
            user.groups.add(3)
            user.save()
            log.info(f'Added user: {email}')

